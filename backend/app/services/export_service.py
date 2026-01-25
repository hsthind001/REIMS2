"""
Export Service - Excel and CSV export functionality

Exports financial data to Excel and CSV formats with proper formatting
"""
from sqlalchemy.orm import Session
from typing import Dict, List, Optional
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.property import Property
from app.models.financial_period import FinancialPeriod
from app.models.balance_sheet_data import BalanceSheetData
from app.models.income_statement_data import IncomeStatementData
from app.models.cash_flow_data import CashFlowData
from app.models.rent_roll_data import RentRollData


class ExportService:
    """Service for exporting financial data to various formats"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def export_balance_sheet_excel(
        self,
        property_code: str,
        year: int,
        month: int,
        organization_id: Optional[int] = None
    ) -> bytes:
        """
        Export balance sheet to Excel with formatting
        
        Returns:
            bytes: Excel file content
        """
        # Get data
        query = self.db.query(Property).filter(Property.property_code == property_code)
        if organization_id is not None:
            query = query.filter(Property.organization_id == organization_id)
        property_obj = query.first()
        
        if not property_obj:
            raise ValueError(f"Property {property_code} not found")
        
        period = self.db.query(FinancialPeriod).filter(
            FinancialPeriod.property_id == property_obj.id,
            FinancialPeriod.period_year == year,
            FinancialPeriod.period_month == month
        ).first()
        
        if not period:
            raise ValueError(f"Period {year}-{month} not found for property {property_code}")
        
        # Get balance sheet data
        bs_data = self.db.query(BalanceSheetData).filter(
            BalanceSheetData.property_id == property_obj.id,
            BalanceSheetData.period_id == period.id
        ).order_by(BalanceSheetData.account_code).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        total_font = Font(bold=True)
        
        # Header
        ws['A1'] = property_obj.property_name
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "Balance Sheet"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"As of {period.period_year}-{period.period_month:02d}-{period.period_end_date.day:02d}"
        ws['A3'].font = Font(size=11)
        
        # Column headers
        row = 5
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "Account"
        ws[f'D{row}'] = "Amount"
        for col in ['A', 'D']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
        
        # Data
        row = 6
        current_section = None
        
        for item in bs_data:
            # Check if this is a section header
            is_section = item.account_code in ['0100-0000', '2000-0000', '3000-0000']
            is_total = item.is_calculated
            
            ws[f'A{row}'] = item.account_code
            ws[f'B{row}'] = item.account_name
            ws[f'D{row}'] = float(item.amount)
            ws[f'D{row}'].number_format = '$#,##0.00'
            
            if is_section or is_total:
                ws[f'B{row}'].font = section_font
                ws[f'D{row}'].font = total_font
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['D'].width = 18
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_income_statement_excel(
        self,
        property_code: str,
        year: int,
        month: int,
        organization_id: Optional[int] = None
    ) -> bytes:
        """Export income statement to Excel"""
        # Similar to balance sheet but with multiple columns (Period, YTD, %)
        query = self.db.query(Property).filter(Property.property_code == property_code)
        if organization_id is not None:
            query = query.filter(Property.organization_id == organization_id)
        property_obj = query.first()
        
        period = self.db.query(FinancialPeriod).filter(
            FinancialPeriod.property_id == property_obj.id,
            FinancialPeriod.period_year == year,
            FinancialPeriod.period_month == month
        ).first()
        
        is_data = self.db.query(IncomeStatementData).filter(
            IncomeStatementData.property_id == property_obj.id,
            IncomeStatementData.period_id == period.id
        ).order_by(IncomeStatementData.account_code).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        
        # Headers
        ws['A1'] = property_obj.property_name
        ws['A2'] = "Income Statement"
        ws['A3'] = f"Period: {year}-{month:02d}"
        
        row = 5
        ws['A5'] = "Account"
        ws['B5'] = "Period"
        ws['C5'] = "YTD"
        ws['D5'] = "Period %"
        ws['E5'] = "YTD %"
        
        # Data
        row = 6
        for item in is_data:
            ws[f'A{row}'] = f"{item.account_code} - {item.account_name}"
            ws[f'B{row}'] = float(item.period_amount)
            ws[f'B{row}'].number_format = '$#,##0.00'
            
            if item.ytd_amount:
                ws[f'C{row}'] = float(item.ytd_amount)
                ws[f'C{row}'].number_format = '$#,##0.00'
            
            if item.period_percentage:
                ws[f'D{row}'] = float(item.period_percentage) / 100
                ws[f'D{row}'].number_format = '0.00%'
            
            if item.ytd_percentage:
                ws[f'E{row}'] = float(item.ytd_percentage) / 100
                ws[f'E{row}'].number_format = '0.00%'
            
            row += 1
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_to_csv(
        self,
        property_code: str,
        year: int,
        month: int,
        document_type: str,
        organization_id: Optional[int] = None
    ) -> bytes:
        """
        Export financial data to CSV
        
        Args:
            property_code: Property code
            year: Period year
            month: Period month
            document_type: Type of document (balance_sheet, income_statement, etc.)
        
        Returns:
            bytes: CSV file content
        """
        query = self.db.query(Property).filter(Property.property_code == property_code)
        if organization_id is not None:
            query = query.filter(Property.organization_id == organization_id)
        property_obj = query.first()
        
        period = self.db.query(FinancialPeriod).filter(
            FinancialPeriod.property_id == property_obj.id,
            FinancialPeriod.period_year == year,
            FinancialPeriod.period_month == month
        ).first()
        
        # Get data based on type
        if document_type == 'balance_sheet':
            data = self.db.query(BalanceSheetData).filter(
                BalanceSheetData.property_id == property_obj.id,
                BalanceSheetData.period_id == period.id
            ).all()
            headers = ['Account Code', 'Account Name', 'Amount', 'Is Calculated']
            rows = [[d.account_code, d.account_name, float(d.amount), d.is_calculated] for d in data]
        
        elif document_type == 'income_statement':
            data = self.db.query(IncomeStatementData).filter(
                IncomeStatementData.property_id == property_obj.id,
                IncomeStatementData.period_id == period.id
            ).all()
            headers = ['Account Code', 'Account Name', 'Period Amount', 'YTD Amount', 'Period %', 'YTD %']
            rows = [[
                d.account_code,
                d.account_name,
                float(d.period_amount),
                float(d.ytd_amount) if d.ytd_amount else None,
                float(d.period_percentage) if d.period_percentage else None,
                float(d.ytd_percentage) if d.ytd_percentage else None
            ] for d in data]
        
        else:
            raise ValueError(f"Unsupported document type: {document_type}")
        
        # Create CSV
        output = BytesIO()
        output.write(b'\xef\xbb\xbf')  # UTF-8 BOM
        writer = csv.writer(output.buffer if hasattr(output, 'buffer') else output, encoding='utf-8')
        
        # Metadata
        writer.writerow([f'Property: {property_obj.property_name} ({property_code})'])
        writer.writerow([f'Period: {year}-{month:02d}'])
        writer.writerow([])
        
        # Headers and data
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        
        output.seek(0)
        return output.getvalue()
