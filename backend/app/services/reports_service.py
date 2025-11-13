"""
ReportsService - Financial Reporting and Excel Export

Generates comprehensive financial reports from database views
Supports summary reports, period comparisons, trend analysis, and Excel exports
"""
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Dict, List, Optional, Any
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


class ReportsService:
    """
    Service for generating financial reports
    
    Features:
    - Complete financial summaries (BS + IS + CF + RR + Metrics)
    - Period-over-period comparisons
    - Annual trend analysis
    - Excel export with professional formatting
    """
    
    def __init__(self, db: Session):
        self.db = db
    
    def get_financial_summary(
        self,
        property_code: str,
        year: int,
        month: int
    ) -> Dict[str, Any]:
        """
        Get complete financial summary for a property/period
        
        Returns all financial data in organized structure:
        - Balance Sheet metrics
        - Income Statement metrics
        - Cash Flow metrics
        - Rent Roll metrics
        - All 35 calculated KPIs
        
        Args:
            property_code: Property code
            year: Period year
            month: Period month
        
        Returns:
            Complete financial summary as nested dict
        """
        # Query the property financial summary view
        query = text("""
            SELECT *
            FROM v_property_financial_summary
            WHERE property_code = :property_code
              AND period_year = :year
              AND period_month = :month
        """)
        
        result = self.db.execute(
            query,
            {"property_code": property_code, "year": year, "month": month}
        ).fetchone()
        
        if not result:
            raise ValueError(f"No financial data found for {property_code} {year}-{month:02d}")
        
        # Convert to dict
        row_dict = dict(result._mapping)
        
        # Debug: Check if we have the new fields
        print(f"DEBUG: total_cash_position={row_dict.get('total_cash_position')}, debt_to_assets_ratio={row_dict.get('debt_to_assets_ratio')}")
        
        # Organize into logical sections
        summary = {
            "property": {
                "property_code": row_dict["property_code"],
                "property_name": row_dict["property_name"],
                "property_type": row_dict["property_type"],
                "city": row_dict["city"],
                "state": row_dict["state"],
            },
            "period": {
                "year": row_dict["period_year"],
                "month": row_dict["period_month"],
                "start_date": str(row_dict["period_start_date"]) if row_dict.get("period_start_date") else None,
                "end_date": str(row_dict["period_end_date"]) if row_dict.get("period_end_date") else None,
            },
            "balance_sheet": {
                "total_assets": self._to_float(row_dict.get("total_assets")),
                "total_liabilities": self._to_float(row_dict.get("total_liabilities")),
                "total_equity": self._to_float(row_dict.get("total_equity")),
                "current_ratio": self._to_float(row_dict.get("current_ratio")),
                "debt_to_equity_ratio": self._to_float(row_dict.get("debt_to_equity_ratio")),
                "debt_to_assets_ratio": self._to_float(row_dict.get("debt_to_assets_ratio")),
                "ltv_ratio": self._to_float(row_dict.get("ltv_ratio")),
                "cash_position": self._to_float(row_dict.get("total_cash_position")),
                "operating_cash": self._to_float(row_dict.get("operating_cash")),
                "restricted_cash": self._to_float(row_dict.get("restricted_cash")),
            },
            "income_statement": {
                "total_revenue": self._to_float(row_dict.get("total_revenue")),
                "total_expenses": self._to_float(row_dict.get("total_expenses")),
                "net_operating_income": self._to_float(row_dict.get("net_operating_income")),
                "net_income": self._to_float(row_dict.get("net_income")),
                "operating_margin": self._to_float(row_dict.get("operating_margin")),
                "profit_margin": self._to_float(row_dict.get("profit_margin")),
            },
            "cash_flow": {
                "operating_cash_flow": self._to_float(row_dict.get("operating_cash_flow")),
                "investing_cash_flow": self._to_float(row_dict.get("investing_cash_flow")),
                "financing_cash_flow": self._to_float(row_dict.get("financing_cash_flow")),
                "net_cash_flow": self._to_float(row_dict.get("net_cash_flow")),
                "beginning_cash_balance": self._to_float(row_dict.get("beginning_cash_balance")),
                "ending_cash_balance": self._to_float(row_dict.get("ending_cash_balance")),
            },
            "rent_roll": {
                "total_units": row_dict.get("total_units"),
                "occupied_units": row_dict.get("occupied_units"),
                "vacant_units": row_dict.get("vacant_units"),
                "occupancy_rate": self._to_float(row_dict.get("occupancy_rate")),
                "total_annual_rent": self._to_float(row_dict.get("total_annual_rent")),
                "avg_rent_per_sqft": self._to_float(row_dict.get("avg_rent_per_sqft")),
            },
            "performance": {
                "noi_per_sqft": self._to_float(row_dict.get("noi_per_sqft")),
                "revenue_per_sqft": self._to_float(row_dict.get("revenue_per_sqft")),
                "expense_ratio": self._to_float(row_dict.get("expense_ratio")),
            },
            "metadata": {
                "calculated_at": str(row_dict["calculated_at"]) if row_dict.get("calculated_at") else None,
            }
        }
        
        return summary
    
    def get_period_comparison(
        self,
        property_code: str,
        start_year: int,
        start_month: int,
        end_year: int,
        end_month: int,
        account_codes: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Compare financial data between two periods
        
        Returns month-over-month changes with variances
        
        Args:
            property_code: Property code
            start_year, start_month: Starting period
            end_year, end_month: Ending period
            account_codes: Optional filter for specific accounts
        
        Returns:
            Comparison data with current, previous, and variance
        """
        # Build query with optional account filter
        where_clause = """
            WHERE property_code = :property_code
              AND period_year = :year
              AND period_month = :month
        """
        
        if account_codes:
            placeholders = ','.join([f":code_{i}" for i in range(len(account_codes))])
            where_clause += f" AND account_code IN ({placeholders})"
        
        query = text(f"""
            SELECT 
                account_code,
                account_name,
                current_month,
                previous_month,
                month_variance,
                variance_percentage,
                ytd_amount,
                is_income
            FROM v_monthly_comparison
            {where_clause}
            ORDER BY account_code
        """)
        
        # Build parameters
        params = {
            "property_code": property_code,
            "year": end_year,
            "month": end_month
        }
        
        if account_codes:
            for i, code in enumerate(account_codes):
                params[f"code_{i}"] = code
        
        results = self.db.execute(query, params).fetchall()
        
        # Format results
        comparison_items = []
        for row in results:
            row_dict = dict(row._mapping)
            comparison_items.append({
                "account_code": row_dict["account_code"],
                "account_name": row_dict["account_name"],
                "current_period": {
                    "year": end_year,
                    "month": end_month,
                    "amount": self._to_float(row_dict["current_month"]),
                },
                "previous_period": {
                    "amount": self._to_float(row_dict["previous_month"]),
                },
                "variance": {
                    "amount": self._to_float(row_dict["month_variance"]),
                    "percentage": self._to_float(row_dict["variance_percentage"]),
                },
                "ytd_amount": self._to_float(row_dict["ytd_amount"]),
                "is_income": row_dict["is_income"],
            })
        
        return {
            "property_code": property_code,
            "start_period": {"year": start_year, "month": start_month},
            "end_period": {"year": end_year, "month": end_month},
            "accounts": comparison_items,
            "total_accounts": len(comparison_items)
        }
    
    def get_annual_trends(
        self,
        property_code: str,
        year: int,
        account_codes: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Get 12-month trends for specific accounts
        
        Returns time-series data suitable for charting
        
        Args:
            property_code: Property code
            year: Year to analyze
            account_codes: List of account codes to include (optional)
        
        Returns:
            Trend data with monthly values for each account
        """
        # Build query
        where_clause = """
            WHERE property_code = :property_code
              AND year = :year
        """
        
        if account_codes:
            placeholders = ','.join([f":code_{i}" for i in range(len(account_codes))])
            where_clause += f" AND account_code IN ({placeholders})"
        
        query = text(f"""
            SELECT
                account_code,
                account_name,
                monthly_amounts,
                monthly_ytd,
                months,
                annual_total,
                monthly_average,
                min_month,
                max_month,
                std_deviation
            FROM v_annual_trends
            {where_clause}
            ORDER BY account_code
        """)
        
        # Build parameters
        params = {"property_code": property_code, "year": year}
        
        if account_codes:
            for i, code in enumerate(account_codes):
                params[f"code_{i}"] = code
        
        results = self.db.execute(query, params).fetchall()
        
        # Format results
        trends = []
        for row in results:
            row_dict = dict(row._mapping)
            
            # Convert arrays to lists
            monthly_amounts = row_dict["monthly_amounts"] if row_dict["monthly_amounts"] else []
            monthly_ytd = row_dict["monthly_ytd"] if row_dict["monthly_ytd"] else []
            months = row_dict["months"] if row_dict["months"] else []
            
            # Convert Decimals to floats
            monthly_amounts = [self._to_float(amt) for amt in monthly_amounts]
            monthly_ytd = [self._to_float(amt) for amt in monthly_ytd]
            
            trends.append({
                "account_code": row_dict["account_code"],
                "account_name": row_dict["account_name"],
                "monthly_data": [
                    {
                        "month": month,
                        "period_amount": amount,
                        "ytd_amount": ytd
                    }
                    for month, amount, ytd in zip(months, monthly_amounts, monthly_ytd)
                ],
                "statistics": {
                    "annual_total": self._to_float(row_dict["annual_total"]),
                    "monthly_average": self._to_float(row_dict["monthly_average"]),
                    "min_month": self._to_float(row_dict["min_month"]),
                    "max_month": self._to_float(row_dict["max_month"]),
                    "std_deviation": self._to_float(row_dict["std_deviation"]),
                }
            })
        
        return {
            "property_code": property_code,
            "year": year,
            "trends": trends,
            "total_accounts": len(trends)
        }
    
    def export_to_excel(
        self,
        property_code: str,
        year: int,
        month: int
    ) -> BytesIO:
        """
        Export financial summary to Excel with professional formatting
        
        Creates multi-sheet workbook:
        - Summary: Key metrics
        - Balance Sheet: Assets, liabilities, equity
        - Income Statement: Revenue, expenses, margins
        - Cash Flow: Operating, investing, financing
        - Rent Roll: Units, occupancy, rent totals
        
        Args:
            property_code: Property code
            year: Period year
            month: Period month
        
        Returns:
            BytesIO stream of Excel file
        """
        # Get financial summary data
        summary = self.get_financial_summary(property_code, year, month)
        
        # Create workbook
        wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(wb, summary)
        self._create_balance_sheet_sheet(wb, summary)
        self._create_income_statement_sheet(wb, summary)
        self._create_cash_flow_sheet(wb, summary)
        self._create_rent_roll_sheet(wb, summary)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    # ==================== EXCEL SHEET CREATORS ====================
    
    def _create_summary_sheet(self, wb: Workbook, summary: Dict):
        """Create Summary sheet with key metrics"""
        ws = wb.active
        ws.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws['A1'] = f"{summary['property']['property_name']} - Financial Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Period: {summary['period']['year']}-{summary['period']['month']:02d}"
        
        # Key metrics table
        row = 4
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        
        row += 1
        metrics_data = [
            ("Total Assets", summary['balance_sheet']['total_assets']),
            ("Total Liabilities", summary['balance_sheet']['total_liabilities']),
            ("Total Equity", summary['balance_sheet']['total_equity']),
            ("", ""),
            ("Total Revenue", summary['income_statement']['total_revenue']),
            ("Total Expenses", summary['income_statement']['total_expenses']),
            ("Net Operating Income", summary['income_statement']['net_operating_income']),
            ("Net Income", summary['income_statement']['net_income']),
            ("", ""),
            ("Operating Margin %", summary['income_statement']['operating_margin']),
            ("Profit Margin %", summary['income_statement']['profit_margin']),
            ("", ""),
            ("Total Units", summary['rent_roll']['total_units']),
            ("Occupied Units", summary['rent_roll']['occupied_units']),
            ("Occupancy Rate %", summary['rent_roll']['occupancy_rate']),
        ]
        
        for label, value in metrics_data:
            ws[f'A{row}'] = label
            if value is not None:
                ws[f'B{row}'] = value
                if isinstance(value, (int, float)) and value >= 1000:
                    ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_balance_sheet_sheet(self, wb: Workbook, summary: Dict):
        """Create Balance Sheet sheet"""
        ws = wb.create_sheet("Balance Sheet")
        
        # Header
        ws['A1'] = "Balance Sheet"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"{summary['property']['property_name']} - {summary['period']['year']}-{summary['period']['month']:02d}"
        
        row = 4
        bs_data = [
            ("Account", "Amount"),
            ("ASSETS", ""),
            ("Total Assets", summary['balance_sheet']['total_assets']),
            ("", ""),
            ("LIABILITIES", ""),
            ("Total Liabilities", summary['balance_sheet']['total_liabilities']),
            ("", ""),
            ("EQUITY", ""),
            ("Total Equity", summary['balance_sheet']['total_equity']),
            ("", ""),
            ("RATIOS", ""),
            ("Current Ratio", summary['balance_sheet']['current_ratio']),
            ("Debt-to-Equity Ratio", summary['balance_sheet']['debt_to_equity_ratio']),
        ]
        
        for label, value in bs_data:
            ws[f'A{row}'] = label
            if value != "":
                ws[f'B{row}'] = value
                if isinstance(value, (int, float)):
                    ws[f'B{row}'].number_format = '#,##0.00'
            
            # Bold section headers
            if label in ["ASSETS", "LIABILITIES", "EQUITY", "RATIOS"]:
                ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_income_statement_sheet(self, wb: Workbook, summary: Dict):
        """Create Income Statement sheet"""
        ws = wb.create_sheet("Income Statement")
        
        ws['A1'] = "Income Statement"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"{summary['property']['property_name']} - {summary['period']['year']}-{summary['period']['month']:02d}"
        
        row = 4
        is_data = [
            ("Account", "Amount"),
            ("REVENUE", ""),
            ("Total Revenue", summary['income_statement']['total_revenue']),
            ("", ""),
            ("EXPENSES", ""),
            ("Total Expenses", summary['income_statement']['total_expenses']),
            ("", ""),
            ("NET OPERATING INCOME", summary['income_statement']['net_operating_income']),
            ("NET INCOME", summary['income_statement']['net_income']),
            ("", ""),
            ("MARGINS", ""),
            ("Operating Margin %", summary['income_statement']['operating_margin']),
            ("Profit Margin %", summary['income_statement']['profit_margin']),
        ]
        
        for label, value in is_data:
            ws[f'A{row}'] = label
            if value != "":
                ws[f'B{row}'] = value
                if isinstance(value, (int, float)):
                    ws[f'B{row}'].number_format = '#,##0.00'
            
            if label in ["REVENUE", "EXPENSES", "MARGINS", "NET OPERATING INCOME", "NET INCOME"]:
                ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_cash_flow_sheet(self, wb: Workbook, summary: Dict):
        """Create Cash Flow sheet with Template v1.0 compliance"""
        from openpyxl.styles import Alignment
        
        ws = wb.create_sheet("Cash Flow")
        
        ws['A1'] = "Cash Flow Statement"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Get cash flow data
        cf = summary.get('cash_flow', {})
        
        row = 3
        
        # Header Section
        ws[f'A{row}'] = "Period:"
        ws[f'B{row}'] = f"{summary['period']['year']}-{summary['period']['month']:02d}"
        row += 1
        ws[f'A{row}'] = "Property:"
        ws[f'B{row}'] = summary['property']['property_code']
        row += 2
        
        # Income Section
        ws[f'A{row}'] = "INCOME"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        income_data = [
            ("Total Income", cf.get('total_income')),
            ("  Base Rentals", cf.get('base_rentals')),
        ]
        
        for label, value in income_data:
            ws[f'A{row}'] = label
            if value is not None:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        row += 1
        
        # Expenses Section
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        expense_data = [
            ("Total Operating Expenses", cf.get('total_operating_expenses')),
            ("Total Additional Expenses", cf.get('total_additional_expenses')),
            ("Total Expenses", cf.get('total_expenses')),
        ]
        
        for label, value in expense_data:
            ws[f'A{row}'] = label
            if value is not None:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        row += 1
        
        # Performance Metrics
        ws[f'A{row}'] = "PERFORMANCE METRICS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        performance_data = [
            ("Net Operating Income (NOI)", cf.get('net_operating_income')),
            ("  NOI %", cf.get('noi_percentage')),
            ("Mortgage Interest", cf.get('mortgage_interest')),
            ("Depreciation", cf.get('depreciation')),
            ("Amortization", cf.get('amortization')),
            ("Net Income", cf.get('net_income')),
            ("  Net Income %", cf.get('net_income_percentage')),
        ]
        
        for label, value in performance_data:
            ws[f'A{row}'] = label
            if value is not None:
                ws[f'B{row}'] = value
                if '%' in label:
                    ws[f'B{row}'].number_format = '0.00"%"'
                else:
                    ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        row += 1
        
        # Cash Flow Summary
        ws[f'A{row}'] = "CASH FLOW SUMMARY"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        cf_summary = [
            ("Net Income", cf.get('net_income')),
            ("Total Adjustments", cf.get('total_adjustments')),
            ("CASH FLOW", cf.get('cash_flow')),
            ("  Cash Flow %", cf.get('cash_flow_percentage')),
            ("", ""),
            ("Beginning Cash Balance", cf.get('beginning_cash_balance')),
            ("Ending Cash Balance", cf.get('ending_cash_balance')),
        ]
        
        for label, value in cf_summary:
            ws[f'A{row}'] = label
            if value is not None and value != "":
                ws[f'B{row}'] = value
                if '%' in label:
                    ws[f'B{row}'].number_format = '0.00"%"'
                else:
                    ws[f'B{row}'].number_format = '#,##0.00'
            
            if label == "CASH FLOW":
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
            
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def _create_rent_roll_sheet(self, wb: Workbook, summary: Dict):
        """Create Rent Roll sheet"""
        ws = wb.create_sheet("Rent Roll")
        
        ws['A1'] = "Rent Roll Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 4
        rr_data = [
            ("Metric", "Value"),
            ("Total Units", summary['rent_roll']['total_units']),
            ("Occupied Units", summary['rent_roll']['occupied_units']),
            ("Vacant Units", summary['rent_roll']['vacant_units']),
            ("Occupancy Rate %", summary['rent_roll']['occupancy_rate']),
            ("", ""),
            ("Total Annual Rent", summary['rent_roll']['total_annual_rent']),
            ("Avg Rent per Sqft", summary['rent_roll']['avg_rent_per_sqft']),
        ]
        
        for label, value in rr_data:
            ws[f'A{row}'] = label
            if value is not None and value != "":
                ws[f'B{row}'] = value
                if isinstance(value, (int, float)) and value >= 1000:
                    ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    # ==================== HELPER METHODS ====================
    
    def _to_float(self, value: Any) -> Optional[float]:
        """Convert Decimal or other numeric types to float for JSON"""
        if value is None:
            return None
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (int, float)):
            return float(value)
        return value

